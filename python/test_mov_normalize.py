# From Python
# It requires OpenCV installed for Python
import sys
import cv2
import os
from sys import platform
import argparse
import time
import numpy as np
import matplotlib.pyplot as plt
import matplotlib
from PIL import Image, ImageDraw, ImageFont
import openpyxl
import math


def write_list_2d(sheet, l_2d, start_row, start_col):
    row_count = 0
    col_count = 0 

    for y, row_data in enumerate(l_2d): #各keypointを列に指定
        # for x, cell_data in enumerate(row_data):
            if col_count == 25:
                row_count += 1
                col_count = 0
            sheet.cell(row=start_row + row_count, column=start_col + col_count, value=l_2d[y])
            col_count += 1
    # for y, row_data in enumerate(l_2d): #1列にすべて表示
    #     # for x, cell_data in enumerate(row_data):
    #         sheet.cell(row=start_row + y, column=start_col, value=l_2d[y])

    

try:
    # Import Openpose (Windows/Ubuntu/OSX)
    dir_path = os.path.dirname(os.path.realpath(__file__))
    try:
        # Change these variables to point to the correct folder (Release/x64 etc.)
        sys.path.append(dir_path + '/../bin/python/openpose/Release')
        os.environ['PATH']  = os.environ['PATH'] + ';' + dir_path + '/../x64/Release;' +  dir_path + '/../bin;'
        import pyopenpose as op
    except ImportError as e:
        print('Error: OpenPose library could not be found. Did you enable `BUILD_PYTHON` in CMake and have this Python script in the right folder?')
        raise e

    # Flags
    parser = argparse.ArgumentParser()
    # parser.add_argument("--image_dir", default="../examples/person/", help="Process a directory of images. Read all standard formats (jpg, png, bmp, etc.).")
    parser.add_argument("--no_display", default=False, help="Enable to disable the visual display.")
    args = parser.parse_known_args()

    # Custom Params (refer to include/openpose/flags.hpp for more parameters)
    params = dict()
    params["model_folder"] = "../models/"

    moviefile = "ichiro2_back.mov"
    # params["render_pose"] = 2  #レンダーにGPUを使用する
    # params["ip_camera"] = 'rtsp://root:P7CEqNF5ui8e@10.5.5.145:554/live1s1.sdp'
    params["video"] = "../examples/movie/" + moviefile  #ファイルパス

    # params["video"] = "../examples/movie/uemura.MOV"  #ファイルパス
    # params["render_threshold"] = 0.05  #レンダーするキーポイントのスレッショルド
    params["write_video"] = "../examples/movie/output_" + moviefile
    # params["net_resolution"] = "-1x512"
    params["number_people_max"] = 1
    params["write_json"] = "../examples/json"
    # params["maximize_positives"] = True  #偽陽性と真陽性の両方を高度に増加させるため、平均精度を害する恐れがある




    # Add others in path?
    for i in range(0, len(args[1])):
        curr_item = args[1][i]
        if i != len(args[1])-1: next_item = args[1][i+1]
        else: next_item = "1"
        if "--" in curr_item and "--" in next_item:
            key = curr_item.replace('-','')
            if key not in params:  params[key] = "1"
        elif "--" in curr_item and "--" not in next_item:
            key = curr_item.replace('-','')
            if key not in params: params[key] = next_item

    # Construct it from system arguments
    # op.init_argv(args[1])
    # oppython = op.OpenposePython()

    # Starting OpenPose
    opWrapper = op.WrapperPython(op.ThreadManagerMode.AsynchronousOut)

    opWrapper.configure(params)
    opWrapper.start()

    counter = 1
    # Main loop
    userWantsToExit = False
    list_output = []
    while not userWantsToExit:
        # list_output.append("特徴量計算場所")
        # list_output.append(str(counter) + "枚目")
        print("特徴量計算場所")

        print(counter, "枚目")
        counter += 1
        # Pop frame
        if counter !=  2700:
            datumProcessed = op.VectorDatum()
            if opWrapper.waitAndPop(datumProcessed):
                if not args[0].no_display:
                    datum = datumProcessed[0]
                    
                    imageToProcess = datum.cvInputData

                    #首と腰のキーポイントを利用して画像をリサイズする。
                    keypoint_chest_x = int(datum.poseKeypoints[0][1][0])
                    keypoint_chest_y = int(datum.poseKeypoints[0][1][1])

                    keypoint_hip_x = int(datum.poseKeypoints[0][8][0])
                    keypoint_hip_y = int(datum.poseKeypoints[0][8][1])

                    sekitui = math.sqrt((keypoint_hip_x - keypoint_chest_x) ** 2  + (keypoint_hip_y - keypoint_chest_y) ** 2)

                    threshold = 80.0 #胸から腰までを80pxにするという基準値

                    if sekitui != 0:
                        resize_rate = round(threshold / sekitui, 2)
                        img_resize = cv2.resize(imageToProcess, (int(imageToProcess.shape[1] * resize_rate), int(imageToProcess.shape[0] * resize_rate)))

                        datum.cvInputData = img_resize
                        keypoint_array = [] #x, y
                        
                        for i in range(25):
                            prob = int(datum.poseKeypoints[0][i][2] * 100)

                            width_r =10 #半径
                            height_r =10 #半径

                            if(prob == 0):
                                keypoint_array.append([-1, -1])
                                print(-10000000)

                                list_output.append(-10000000)

                            else:
                                keypoint_x = int(datum.poseKeypoints[0][i][0] * resize_rate)
                                keypoint_y = int(datum.poseKeypoints[0][i][1] * resize_rate)
                                keypoint_array.append([keypoint_x, keypoint_y])
                                img_area = img_resize[keypoint_y - height_r : keypoint_y + height_r, keypoint_x - width_r : keypoint_x + width_r]

                                rsum,gsum,bsum = 0.0,0.0,0.0
                                

                                #画像の範囲指定をしたimg_areaからRGBをそれぞれ2次元配列で全px分取得。それをravelで1次元化してから平均を取得している
                                ravg = np.ravel(img_area[:, :, 2]).mean() 
                                gavg = np.ravel(img_area[:, :, 1]).mean()
                                bavg = np.ravel(img_area[:, :, 0]).mean()

                                # hsv = cv2.cvtColor(np.array([[[bavg, gavg, ravg]]], dtype=np.uint8), cv2.COLOR_BGR2HSV)[0][0]
                                # print("H:", hsv[0])
                                # list_output.append(hsv[0])
                                list_output.append(bavg)




                    cv2.imshow("OpenPose 1.7.0 - Tutorial Python API", datum.cvOutputData)
                    # cv2.imshow("OpenPose 1.7.0 - Tutorial Python API", HSV_img)


                    key = cv2.waitKey(1)
                    key == 27
            else:
                break
        else:
            break;    

    # excel_path = r"C:\Users\isapo\openpose-1.7.0-binaries-win64-gpu-python3.7-flir-3d_recommended\openpose\examples\20x20.xlsx"
    # excel_path = r"C:\Users\isapo\openpose-1.7.0-binaries-win64-gpu-python3.7-flir-3d_recommended\openpose\examples\distance.xlsx"
    excel_path = r"C:\Users\isapo\openpose-1.7.0-binaries-win64-gpu-python3.7-flir-3d_recommended\openpose\examples\20x20_RGB.xlsx"


    wb = openpyxl.load_workbook(excel_path)  
    sheet = wb.worksheets[6]

    write_list_2d(sheet, list_output, 3,2) #行、列

    wb.save(excel_path)
    wb.close()

except Exception as e:
    print(e)
    sys.exit(-1)
