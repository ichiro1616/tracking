# From Python
# It requires OpenCV installed for Python
import sys
import cv2
import os
from sys import platform
import argparse
import time
import numpy as np
import matplotlib.pyplot as plt
import matplotlib
from PIL import Image, ImageDraw, ImageFont
import openpyxl
import math



def write_list_2d(sheet, l_2d, start_row, start_col):
    row_count = 0
    col_count = 0 

    for y, row_data in enumerate(l_2d): #各keypointを列に指定
        # for x, cell_data in enumerate(row_data):
            if col_count == 9:
                row_count += 1
                col_count = 0
            sheet.cell(row=start_row + row_count, column=start_col + col_count, value=l_2d[y])
            col_count += 1


    # for y, row_data in enumerate(l_2d): #1列にすべて表示
    #     # for x, cell_data in enumerate(row_data):
    #         sheet.cell(row=start_row + y, column=start_col, value=l_2d[y])

    



def sim_distance(person_now, person_previous): #0フレーム目の人と色相を比較し類似度を算出
    sum_of_sqrt = 0
    for x in range(len(person_now)):
        if person_now[x] != -1 or person_previous != -1:
            calculation_source =  abs(int(person_now[x]) - int(person_previous[x]))
            ring_calculate = abs(calculation_source - 180) if round((calculation_source / 180)) * 180 > 90 else calculation_source
            # print(calculation_source, round((calculation_source / 180)) * 180, ring_calculate)
            sum_of_sqrt += (ring_calculate)**2
    sum_of_sqrt = math.sqrt(sum_of_sqrt)
    bottom = math.sqrt((90)**2 * len(person_now)) #色相の差分の最大値の2乗xキーポイントの数の平方根をとっている
    # print(sum_of_sqrt, 1 - sum_of_sqrt/bottom)
    return 1 - sum_of_sqrt/ bottom

def list_tranpose(matrix): #2次元配列を転置している
    # print(type(matrix))
    # <class 'numpy.ndarray'>
    matrix_np_t = matrix.T
    # print(matrix_np_t)

    matrix_np_t_list = matrix_np_t.tolist()
    # print(type(matrix_np_t_list))
    # <class 'list'>
    return matrix_np_t_list

def search_minmax(list_tr): #信頼度が一定以上のキーポイントの中で最小最大の値をxy座標ごとに返す
    keypoint_x_array = list_tr[0] #人物を四角で囲うために全キーポイントでの最小、最大のxy座標を取得
    keypoint_y_array = list_tr[1]
    reli_array = list_tr[2]
    for rel in range(len(reli_array)): #各キーポイントで信頼度が一定以下なら、そのxy座標を平均化して最小、最大で取得できないようにする。
        if list_tr[2][rel] < 0.5:
            keypoint_x_array[rel] = np.mean(keypoint_x_array)
            keypoint_y_array[rel] = np.mean(keypoint_y_array)


    keypoint_x_min = min(keypoint_x_array)
    keypoint_y_min = min(keypoint_y_array)
    keypoint_x_max = max(keypoint_x_array)
    keypoint_y_max = max(keypoint_y_array)
    return [keypoint_x_min, keypoint_y_min], [keypoint_x_max, keypoint_y_max]


try:
    # Import Openpose (Windows/Ubuntu/OSX)
    dir_path = os.path.dirname(os.path.realpath(__file__))
    try:
        # Change these variables to point to the correct folder (Release/x64 etc.)
        sys.path.append(dir_path + '/../bin/python/openpose/Release')
        os.environ['PATH']  = os.environ['PATH'] + ';' + dir_path + '/../x64/Release;' +  dir_path + '/../bin;'
        import pyopenpose as op
    except ImportError as e:
        print('Error: OpenPose library could not be found. Did you enable `BUILD_PYTHON` in CMake and have this Python script in the right folder?')
        raise e

    # Flags
    parser = argparse.ArgumentParser()
    # parser.add_argument("--image_dir", default="../examples/person/", help="Process a directory of images. Read all standard formats (jpg, png, bmp, etc.).")
    parser.add_argument("--no_display", default=False, help="Enable to disable the visual display.")
    args = parser.parse_known_args()

    # Custom Params (refer to include/openpose/flags.hpp for more parameters)
    params = dict()
    params["model_folder"] = "../models/"

    moviefile = "pattern1.MOV"
    # params["render_pose"] = 2  #レンダーにGPUを使用する
    # params["ip_camera"] = 'rtsp://root:P7CEqNF5ui8e@10.5.5.145:554/live1s1.sdp'
    params["video"] = "../examples/movies_gym/" + moviefile  #ファイルパス

    # params["render_threshold"] = 0.05  #レンダーするキーポイントのスレッショルド
    params["write_video"] = "../examples/movies_gym/output_tracking_" + moviefile
    # params["net_resolution"] = "-1x512"
    params["number_people_max"] = 5
    params["write_json"] = "../examples/json"
    # params["maximize_positives"] = True  #偽陽性と真陽性の両方を高度に増加させるため、平均精度を害する恐れがある




    # Add others in path?
    for i in range(0, len(args[1])):
        curr_item = args[1][i]
        if i != len(args[1])-1: next_item = args[1][i+1]
        else: next_item = "1"
        if "--" in curr_item and "--" in next_item:
            key = curr_item.replace('-','')
            if key not in params:  params[key] = "1"
        elif "--" in curr_item and "--" not in next_item:
            key = curr_item.replace('-','')
            if key not in params: params[key] = next_item

    # Construct it from system arguments
    # op.init_argv(args[1])
    # oppython = op.OpenposePython()

    # Starting OpenPose
    opWrapper = op.WrapperPython(op.ThreadManagerMode.AsynchronousOut)

    opWrapper.configure(params)
    opWrapper.start()

    counter = 1
    # Main loop
    userWantsToExit = False
    person_h_array = []
    
    keypoint_judge =  {"鼻" : 0, "胸" : 1, "右肩" : 1, "右ひじ" : 1, "右手首" : 0, "左肩" : 1, "左ひじ" : 1, "左手首" : 0, "腰" : 1, "右腰" : 0, "右ひざ" : 1, "右足首" : 0, "左腰" : 0, "左ひざ" : 1, "左足首" : 0, "右目" : 0, "左目" : 0, "右耳" : 0, "左耳" : 0, "左親指" : 0, "左小指" : 0, "左かかと" : 0, "右親指" : 0, "右小指" : 0, "右かかと" : 0}
    while not userWantsToExit:
        if counter == 200:
            break
        else:

            # person_h_array = []
            person_keypoints_array = []
            person_keypoints_min_array = []
            person_keypoints_max_array = []

            person_box_array = [[255,0,0], [0,255,0], [0,0,255], [255,255,0], [255,0,255],[255,0,255],[0,255,255],[255,255,255],[128,128,255],[255,128,255]] #B,G,Rを人数分
            # h_list_output.append("特徴量計算場所")
            # h_list_output.append(str(counter) + "枚目")
            print("-------------------------------------------------")
            
            # print("特徴量計算場所")

            print(counter, "枚目")


            # Pop frame
            datumProcessed = op.VectorDatum()
            if opWrapper.waitAndPop(datumProcessed):
                if not args[0].no_display:
                    datum = datumProcessed[0]
                    imageToProcess = datum.cvInputData
                
                    h_array = []
                    keypoints_array = []
                    keypoints_min_array = []
                    keypoints_max_array = []
                    

                    for d in range(len(datum.poseKeypoints)): #識別されている人数だけ回す
                        print(str(d + 1) + "人目")
                        keypoints_list_output = []
                        h_list_output = []
                        # print(datum.poseKeypoints)


                        person_h_array.append(counter)



                        for i, val in enumerate(keypoint_judge.values()): #keypointの数だけ回す。余裕があったらkeypoint_judgeの中で1になっている数だけ回す。そのindexを取得するようにしたら、1行下のif文がいらなくなる。
                            if val == 1: #keypointを使うかどうか

                                prob = int(datum.poseKeypoints[d][i][2] * 100)


                                if(prob == 0):
                                    keypoints_list_output.append([-1, -1])
                                    h_list_output.append(-1)
                                    person_h_array.append(-1)
                                    # print("H:", -1)


                                else:

                                    width_r =10 #半径 
                                    height_r =10 #半径
                                    keypoint_x = int(datum.poseKeypoints[d][i][0])
                                    keypoint_y = int(datum.poseKeypoints[d][i][1])
                                    keypoints_list_output.append([keypoint_x, keypoint_y])
                                    img_area = imageToProcess[keypoint_y - height_r : keypoint_y + height_r, keypoint_x - width_r : keypoint_x + width_r]

                                    print("|||||",img_area.shape[1], "||||", img_area.shape[0])

                                    rsum,gsum,bsum = 0.0,0.0,0.0

                                    #画像の範囲指定をしたimg_areaからRGBをそれぞれ2次元配列で全px分取得。それをravelで1次元化してから平均を取得している
                                    ravg = np.ravel(img_area[:, :, 2]).mean() 
                                    gavg = np.ravel(img_area[:, :, 1]).mean()
                                    bavg = np.ravel(img_area[:, :, 0]).mean()
                                    # print(img_area[:, :, 2])
                                    # print(img_area[:, :, 1])
                                    # print(img_area[:, :, 0])
                                    # print(bavg,gavg,ravg)

                                    if math.isnan(bavg):
                                        h_list_output.append(-1)
                                        person_h_array.append(-1)
                                        print("H:NAN")


                                    else:
                                        hsv = cv2.cvtColor(np.array([[[bavg, gavg, ravg]]], dtype=np.uint8), cv2.COLOR_BGR2HSV)[0][0]
                                        print("H:", hsv[0])
                                        h_list_output.append(hsv[0])
                                        person_h_array.append(hsv[0])


                        list_tranpose_array = list_tranpose(datum.poseKeypoints[d]) #キーポイントデータを転置してlistにしている
                        
                        keypoint_minmax = search_minmax(list_tranpose_array)
                        # print(keypoints_list_output)

                        h_array.append(h_list_output)
                        keypoints_array.append(keypoints_list_output)
                        keypoints_min_array.append([int(keypoint_minmax[0][0]), int(keypoint_minmax[0][1])]) #float型をint型にしている
                        keypoints_max_array.append([int(keypoint_minmax[1][0]), int(keypoint_minmax[1][1])])




                # person_h_array = h_array[:]
                person_keypoints_array = keypoints_array[:]
                person_keypoints_min_array = keypoints_min_array[:]
                person_keypoints_max_array = keypoints_max_array[:]

                # print(len(person_h_array), len(person_keypoints_array))
                # print("---",len(person_keypoints_min_array), person_keypoints_min_array)


                # print("現フレーム：",person_h_array)
                # cv2.imshow("OpenPose 1.7.0 - Tutorial Python API", imageToProcess)
                cv2.imshow("OpenPose 1.7.0 - Tutorial Python API", datum.cvOutputData)
                # img_area = cv2.resize(img_area, (1080, 1920))
                # cv2.imshow("OpenPose 1.7.0 - Tutorial Python API", img_area)


                key = cv2.waitKey(1)
                key == 27
                
            else:
                break
            counter += 1

    excel_path = r"C:\Users\isapo\openpose-1.7.0-binaries-win64-gpu-python3.7-flir-3d_recommended\openpose\examples\gym_data.xlsx"
    wb = openpyxl.load_workbook(excel_path)  
    sheet = wb.worksheets[2]

    # write_list_2d(sheet, person_h_array, 2,1) #行、列

    wb.save(excel_path)
    wb.close()

except Exception as e:
    print(e)
    sys.exit(-1)
